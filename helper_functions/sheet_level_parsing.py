from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import cell as xlcell
from consts.days_of_week import weekdays
from consts.lecture_time_ranges import lecture_time_ranges, first_time_range, second_time_range, third_time_range
from consts.lecture_time_ranges import fourth_time_range, fifth_time_range, sixth_time_range, seventh_time_range
from db_api.model.schedule_group_type_day import ScheduleGroupTypeDay
import re


async def parse_worksheet(worksheet: Worksheet, academic_degree, faculty, year, dryrun=True):
    start_column_index, end_column_index, groupname_row_index = get_column_indexes_and_groupname_row_index(worksheet)
    time_range_width, start_row_index, time_ranges_in_a_day = \
        get_time_range_width_and_start_row_index_and_timeranges_count(worksheet)

    for column_index in range(start_column_index, end_column_index + 1):
        await parse_schedule_for_one_group(worksheet=worksheet, column_index=column_index,
                                     groupname_row_index=groupname_row_index, start_row_index=start_row_index,
                                     time_range_width=time_range_width, time_ranges_in_a_day=time_ranges_in_a_day,
                                     academic_degree=academic_degree, faculty=faculty, year=year,
                                     dryrun=dryrun)



async def parse_schedule_for_one_group(worksheet: Worksheet, column_index, groupname_row_index, start_row_index,
                                       time_range_width, time_ranges_in_a_day,
                                       academic_degree, faculty, year, dryrun=True):

    accidental_service_words_or_empty = ('', ' ', '_', '-', None)

    #Расписание на дни недели для недели "ЧИСЛИТЕЛЬ"
    group_even_week_schedules = []
    EVEN_WEEK_TYPE = False #ЧИСЛИТЕЛЬ - 0 - False

    # Расписание на дни недели для недели "ЗНАМЕНАТЕЛЬ"
    group_odd_week_schedules = []
    ODD_WEEK_TYPE = True #ЗНАМЕНАТЕЛЬ - 1 - True

    group_name = str(worksheet.cell(row=groupname_row_index, column=column_index).value)
    if group_name == "None":
        return
    group_name = group_name.strip()
    if len(group_name.split(" ")) > 1:
        group_name = group_name.split(" ")[-1]
    else:
        group_name = group_name.split(".")[-1]

    group_schedules = []

    row_index = start_row_index

    end_row_index = worksheet.max_row

    # day_info[0] is name, [1] - index 1-6
    for day_info in weekdays:
        time_ranges_in_a_day = get_timeranges_count(worksheet, day_info)
        time_range_counter = 1

        time_ranges = lecture_time_ranges
        if time_ranges_in_a_day == 5:
            time_ranges = time_ranges[:-2]

        if time_ranges_in_a_day == 6:
            time_ranges = time_ranges[:-1]

        lecture_num = 1
        for time_range in time_ranges:
            even_lectures = []
            odd_lectures = []
            # 1 time range block length in rows
            for i in range(time_range_width):
                cell_value = get_cell_value(worksheet,
                                            worksheet.cell(row=row_index, column=column_index))
                if cell_value not in accidental_service_words_or_empty:

                    if i < (time_range_width // 2):
                        if str(cell_value).strip() not in even_lectures:
                            even_lectures.append(str(cell_value).strip())
                    else:
                        if str(cell_value).strip() not in odd_lectures:
                            odd_lectures.append(str(cell_value).strip())
                row_index += 1



            if len(even_lectures) > 0:
                even_lecture_full_name = " ".join(even_lectures)

                if(even_lecture_full_name != "-" and even_lecture_full_name.find("=") == -1):
                    schedule_group_day_even = create_day(group_name, even_lecture_full_name, EVEN_WEEK_TYPE,
                                                             day_info[1], lecture_num)
                    if dryrun:
                        print(schedule_group_day_even)
                    else:
                        await schedule_group_day_even.create()
                even_lectures.clear()

            if len(odd_lectures) > 0:
                odd_lecture_full_name = " ".join(odd_lectures)
                if (odd_lecture_full_name != "-" and odd_lecture_full_name.find("=") == -1):
                    schedule_group_day_odd = create_day(group_name, odd_lecture_full_name, ODD_WEEK_TYPE,
                                                            day_info[1], lecture_num)

                    if dryrun:
                        print(schedule_group_day_odd)
                    else:
                        try:
                            await schedule_group_day_odd.create()
                        except Exception as err:
                            print("ERROR: ", str(err), "group:"+str(group_name), "day:"+str(day_info[1]), "num:"+str(lecture_num), "ws:"+":".join([worksheet.title, str(row_index), str(column_index)]))
                odd_lectures.clear()
            lecture_num += 1
            if time_range_counter == time_ranges_in_a_day:
                # if we're in last time range, skip 1 empty row before next day, append schedule_group_day to list
                row_index += 1

            if row_index == end_row_index:
                break

            time_range_counter += 1
            even_lectures.clear()
            odd_lectures.clear()


def create_day(group_name, lecture_full_name, WEEK_TYPE, day_info,number_of_lesson) -> any:
    if lecture_full_name != "Нет занятий.":
                lecture_full_name_without_raplace = lecture_full_name
                lecture_full_name = remove_nextline_symbol(lecture_full_name)
    schedule_group_day = ScheduleGroupTypeDay(
                group_name=group_name,
                week=WEEK_TYPE,
                day_of_week=day_info,
                number_of_less=number_of_lesson,
                classroom=lecture_full_name.split(" ")[-1])
    try:
                class_str = lecture_full_name_without_raplace.split("\n")[-1]
                class_str = class_str.replace("\n", "").rstrip().lstrip()
                if class_str.find("D;") != -1:
                    class_str = class_str.replace(" MS Teams: ", "")
    except ValueError:
                class_str = "-"
    try:
                disc = lecture_full_name_without_raplace[
                          0:lecture_full_name_without_raplace.index("\n", 0, len(lecture_full_name_without_raplace))]
                disc = disc[0:disc.rindex("(", 0, len(disc))]
    except ValueError:
        try:
                    disc = lecture_full_name_without_raplace[
                          0:lecture_full_name_without_raplace.index("\n", 0, len(lecture_full_name_without_raplace))]
        except ValueError:
                    disc = "-"

    try:
                formats_dump = lecture_full_name_without_raplace[0:lecture_full_name_without_raplace.index("\n", 0, len(lecture_full_name_without_raplace))]
                formats = formats_dump[formats_dump.rindex("(", 0, len(formats_dump))+1:formats_dump.rindex(")", 0, len(formats_dump))]
                if not((formats == "Л") or (formats == "пр") or (formats == "лаб") or (formats == "пр/лаб")):
                    formats = ("-")
    except ValueError:
                formats = ("-")

    try:
                professors = lecture_full_name_without_raplace[lecture_full_name_without_raplace.index("\n", 0, len(lecture_full_name_without_raplace)) + 1:lecture_full_name_without_raplace.rindex("\n", 0, len(lecture_full_name_without_raplace))]
                professors = remove_nextline_symbol(professors)
                professors = professors.rstrip().lstrip()
    except ValueError:
                print(lecture_full_name_without_raplace)
                professors = ("-")

    if len(class_str) > 30:
                class_str = class_str[0:29]

    schedule_group_day.professor = professors
    schedule_group_day.format = formats
    schedule_group_day.classroom = class_str
    schedule_group_day.discipline = disc.strip()
    return schedule_group_day


# section get cell value
def within_range(bounds: tuple, cell: xlcell) -> bool:
    column_start, row_start, column_end, row_end = bounds
    row = cell.row
    if row >= row_start and row <= row_end:
        column = cell.column
        if column >= column_start and column <= column_end:
            return True
    return False


def get_cell_value(sheet: Worksheet, cell: xlcell) -> any:
    for merged in sheet.merged_cells:
        if within_range(merged.bounds, cell):
            return sheet.cell(merged.min_row, merged.min_col).value
    return cell.value



# section indexes
def get_column_indexes_and_groupname_row_index(target_worksheet: Worksheet):
    for row in target_worksheet.iter_rows():
        for cell in row:
            if str(cell.value).lower().startswith("время занятий"):
                groupname_row_index = cell.row
                starting_column_index = cell.column + 2
                ending_column_index = target_worksheet.max_column
                return starting_column_index, ending_column_index, groupname_row_index



def get_time_range_width_and_start_row_index_and_timeranges_count(target_worksheet: Worksheet):
    day_start_row = 0
    day_end_row = 0
    start_row_index = 0
    MONDAY_NAMES = ("ПОНЕДЕЛЬНИК", "понедельник", "Понедельник")
    FIFTH_TIME_RANGE = '1645-1815'
    for row in target_worksheet.iter_rows(max_col=1):
        for cell in row:
            cell_value = get_cell_value(target_worksheet, cell)
            if (day_start_row == 0) and (cell_value in MONDAY_NAMES):
                day_start_row = cell.row
                start_row_index = cell.row
                continue
            if (day_start_row != 0) and (cell_value not in MONDAY_NAMES):
                day_end_row = cell.row
                day_width = day_end_row - day_start_row
                time_range = get_cell_value(target_worksheet,
                                            target_worksheet.cell(row=(cell.row - 1), column=(cell.column + 2)))
                if str(time_range) == FIFTH_TIME_RANGE:
                    time_ranges_in_a_day = 5
                else:
                    time_ranges_in_a_day = 7
                time_range_width = day_width // time_ranges_in_a_day
                return time_range_width, start_row_index, time_ranges_in_a_day


def get_timeranges_count(target_worksheet: Worksheet, day_info):
    day_start_row = 0
    day_end_row = 0
    start_row_index = 0
    DAY_NAMES = day_info[0].upper()
    FIFTH_TIME_RANGE = '1645-1815'
    SIXTH_TIME_RANGE = '1830-2000'
    time_ranges_in_a_day = 5
    if DAY_NAMES == "СУББОТА":
        return 5
    for row in target_worksheet.iter_rows(max_col=1):
        for cell in row:
            cell_value = get_cell_value(target_worksheet, cell)
            if (day_start_row == 0) and (cell_value == DAY_NAMES):
                day_start_row = cell.row
                start_row_index = cell.row
                continue
            if (day_start_row != 0) and (cell_value != DAY_NAMES):
                day_end_row = cell.row
                day_width = day_end_row - day_start_row
                time_range = get_cell_value(target_worksheet,
                                            target_worksheet.cell(row=(cell.row - 1), column=(cell.column + 2)))

                if str(time_range) == FIFTH_TIME_RANGE:
                    time_ranges_in_a_day = 5
                elif str(time_range) == SIXTH_TIME_RANGE:
                    time_ranges_in_a_day = 6
                else:
                    time_ranges_in_a_day = 7
                time_range_width = day_width // time_ranges_in_a_day

                return time_ranges_in_a_day

# section utils - fuction to remove nextline symbols
def remove_nextline_symbol(target_string):
    target_string = target_string.replace("\n", " ")
    return target_string




